import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import tempfile
import os

import re, io, unicodedata, statistics
from dataclasses import dataclass, field
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import List, Dict, Tuple, Optional

import fitz
import cv2
import numpy as np
import pytesseract

PARSER_VERSION = "2.3-layout-native"

MONTHS = {
    'GEN':1,'GENNAIO':1,'FEB':2,'FEBBRAIO':2,'MAR':3,'MARZO':3,'APR':4,'APRILE':4,
    'MAG':5,'MAGGIO':5,'GIU':6,'GIUGNO':6,'LUG':7,'LUGLIO':7,'AGO':8,'AGOSTO':8,
    'SET':9,'SETT':9,'SETTEMBRE':9,'OTT':10,'OTTOBRE':10,'NOV':11,'NOVEMBRE':11,'DIC':12,'DICEMBRE':12
}

def clean(s):
    s=(s or '').replace('\u00a0',' ')
    s=unicodedata.normalize('NFKD',s)
    s=''.join(c for c in s if not unicodedata.combining(c))
    s=s.upper().replace('’',"'").replace('`',"'")
    s=re.sub(r'\s+',' ',s).strip(' |\t\r\n')
    return s

def key(s):
    s=clean(s)
    s=re.sub(r'[^A-Z0-9]+','',s)
    return s

def ratio(a,b):
    a,b=key(a),key(b)
    if not a or not b: return 0
    if a==b:return 1.0
    if len(a)>=5 and (a in b or b in a): return 0.96
    return SequenceMatcher(None,a,b).ratio()

def normalize_time(t):
    t=clean(t).replace('.',':')
    m=re.search(r'\b([0-2]?\d):([0-5]\d)\b',t)
    if not m:return ''
    h=int(m.group(1))
    if h>23:return ''
    return f'{h:02d}:{m.group(2)}'

def normalize_numeric_date(s):
    m=re.search(r'\b(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})\b',s)
    if not m:return ''
    d,mn,y=map(int,m.groups())
    if y<100:y+=2000
    try:return datetime(y,mn,d).strftime('%d/%m/%Y')
    except:return ''

def season_years(text):
    m=re.search(r'STAGIONE\s+(\d{4})\s*[-/]\s*(\d{4})', clean(text))
    if m:return int(m.group(1)),int(m.group(2))
    years=[int(x) for x in re.findall(r'20\d{2}',text)]
    if years:
        y=min(years); return y,y+1
    return datetime.now().year, datetime.now().year+1

def date_from_day_month(day, mon, y1,y2, phase=''):
    monu=clean(mon).replace('.','')
    month=MONTHS.get(monu[:3], MONTHS.get(monu))
    if not month:return ''
    # football season: Aug-Dec in first year, Jan-Jun in second year
    year = y1 if month>=7 else y2
    if 'PRIMAVERILE' in clean(phase): year=y2
    try:return datetime(year,month,int(day)).strftime('%d/%m/%Y')
    except:return ''

def parse_textual_dates(line,y1,y2,phase=''):
    # Numeric dates first
    nums=re.findall(r'\b\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\b',line)
    if nums:return [normalize_numeric_date(x) for x in nums]
    vals=[]
    for d,m in re.findall(r'\b(\d{1,2})\s+([A-Za-zÀ-ÿ]{3,10})\b',line):
        dt=date_from_day_month(d,m,y1,y2,phase)
        if dt:vals.append(dt)
    return vals

@dataclass
class TeamInfo:
    name:str
    locality:str=''
    address:str=''
    time:str=''
    field_name:str=''

@dataclass
class Match:
    date:str
    home:str
    away:str
    time:str=''
    locality:str=''
    address:str=''
    round_no:str=''

@dataclass
class Section:
    competition:str
    group:str
    teams:Dict[str,TeamInfo]=field(default_factory=dict)
    matches:List[Match]=field(default_factory=list)
    source_format:str=''

    @property
    def label(self):
        g=f' - Girone {self.group}' if self.group else ''
        return f'{self.competition}{g}'.strip()



def parse_field_table_words(page):
    """Parse LND field tables from positioned PDF words (robust to broken text order)."""
    words=page.get_text('words')
    if not words:return {}
    hdr={}
    for w in words:
        txt=clean(w[4])
        if txt.startswith('SOCIETA') and 'soc' not in hdr: hdr['soc']=w[0]
        elif txt.startswith('CAMPO') and 'code' not in hdr: hdr['code']=w[0]
        elif txt.startswith('DENOMINAZIONE') and 'field' not in hdr: hdr['field']=w[0]
        elif txt.startswith('LOCALITA') and 'loc' not in hdr: hdr['loc']=w[0]
        elif txt=='ORA' and 'time' not in hdr: hdr['time']=w[0]
        elif txt.startswith('INDIRIZZO') and 'addr' not in hdr: hdr['addr']=w[0]
    if not all(k in hdr for k in ['soc','code','field','loc','time','addr']): return {}
    # vertical separator glyphs from the ASCII-style table
    barxs=[]
    for w in words:
        if w[4].strip()=='|':barxs.append(w[0])
    clusters=[]
    for x in sorted(barxs):
        if not clusters or abs(x-clusters[-1][0])>1.5:clusters.append([x,1])
        else:
            clusters[-1][0]=(clusters[-1][0]*clusters[-1][1]+x)/(clusters[-1][1]+1);clusters[-1][1]+=1
    common=[x for x,n in clusters if n>=5]
    def left_bar(x,default):
        vals=[b for b in common if b<x]
        return max(vals) if vals else default
    soc_left=left_bar(hdr['soc'],hdr['soc']-10)
    code_left=left_bar(hdr['code'],hdr['code']-10)
    field_left=left_bar(hdr['field'],hdr['field']-10)
    loc_left=hdr['loc']-1
    time_left=left_bar(hdr['time'],hdr['time']-12)
    addr_left=left_bar(hdr['addr'],hdr['addr']-10)
    # end after address; first common bar to the right, otherwise page edge
    right=[b for b in common if b>addr_left+20]
    addr_right=min(right) if right else page.rect.width+1
    bounds=[soc_left,code_left,field_left,loc_left,time_left,addr_left,addr_right]
    header_y=max(w[1] for w in words if clean(w[4]).startswith(('SOCIETA','DENOMINAZIONE','LOCALITA','INDIRIZZO')))
    data=[w for w in words if w[1]>header_y+4]
    groups=[]
    for w in sorted(data,key=lambda x:((x[1]+x[3])/2,x[0])):
        yc=(w[1]+w[3])/2
        if not groups or abs(yc-groups[-1][0])>2.8:groups.append([yc,[w]])
        else:groups[-1][1].append(w);groups[-1][0]=(groups[-1][0]+yc)/2
    teams={}
    for yc,ws in groups:
        cols=[]
        for a,b in zip(bounds,bounds[1:]):
            arr=sorted([w for w in ws if a+0.2<=w[0]<b-0.2 and w[4].strip()!='|'],key=lambda w:w[0])
            cols.append(clean(' '.join(w[4] for w in arr)))
        name,code,fieldname,locality,tm,addr=cols
        if not name or not re.fullmatch(r'\d{1,5}',code):continue
        if name.startswith('SOCIETA'):continue
        teams[name]=TeamInfo(name,locality,addr,normalize_time(tm),fieldname)
    return teams

# ---------- field table text parser ----------
def parse_header_comp_group(text):
    t=clean(text)
    # Try GIRONE: X
    gm=re.search(r'(.{3,120}?)\s+GIRONE\s*:\s*([A-Z0-9]+)',t)
    if gm:
        comp=gm.group(1)
        comp=re.sub(r'^.*?LOMBARDIA\s*','',comp).strip(' *|-')
        return comp,gm.group(2)
    gm=re.search(r'GIRONE\s+([A-Z0-9]+)',t)
    group=gm.group(1) if gm else ''
    # graphical header lines
    lines=[clean(x) for x in text.splitlines() if clean(x)]
    comp=''
    for ln in lines[:10]:
        if any(w in ln for w in ['CALENDARIO','STAGIONE','GIRONE','COMITATO','LOMBARDIA']): continue
        if len(ln)>3:
            comp=ln; break
    return comp,group

def parse_field_table_text(text):
    teams={}
    lines=text.splitlines()
    # regular aligned rows
    for ln in lines:
        if '|' not in ln: continue
        cells=[c.strip() for c in ln.split('|')]
        # keep internal blanks but strip edge blanks
        while cells and not cells[0]:cells.pop(0)
        while cells and not cells[-1]:cells.pop()
        if len(cells)<5:continue
        name=clean(cells[0])
        code=clean(cells[1])
        if not re.fullmatch(r'\d{1,5}',code):continue
        fieldloc=clean(cells[2])
        tm=normalize_time(cells[3])
        addr=clean(cells[4])
        if not name or name in ["SOCIETA'",'SOCIETA']:continue
        locality=extract_locality_from_field(fieldloc)
        teams[name]=TeamInfo(name,locality,addr,tm,fieldloc)
    return teams

def extract_locality_from_field(s):
    s=clean(s)
    if not s:return ''
    # Common explicit dash separator used in modern graphical tables
    parts=[p.strip() for p in re.split(r'\s+-\s+',s) if p.strip()]
    if len(parts)>=2:
        return parts[-1]
    # remove quoted venue names and parentheticals
    x=re.sub(r'"[^"]+"', ' ', s)
    x=re.sub(r'\([^)]*\)', ' ', x)
    x=re.sub(r'\bE\.?A\.?\b',' ',x)
    # if there is CAMPO N... take suffix
    m=list(re.finditer(r'\bCAMPO\s*(?:N\.?\s*)?[0-9A-Z°.-]*\s*',x))
    if m:
        cand=x[m[-1].end():].strip(' -')
        if cand and not re.match(r'^\d',cand): return cand
    # remove common facility prefixes
    x=re.sub(r'^(?:C\.?S\.?|CENTRO SPORTIVO|CENTRO SPOTIVO|CAMPO SPORTIVO|STADIO|COMUNALE|ORATORIO|PARROCCHIALE|C\.COM\.|C\.S\.COMUNALE)[ .-]*','',x)
    x=re.sub(r'^(?:COMUNALE|SPORTIVO|N\.?\s*\d+|CAMPO\s*\d+)[ .-]*','',x)
    # after final explicit N.1 / N.2 often locality follows
    m=re.search(r'\bN\.?\s*\d+\s+(.+)$',x)
    if m:return m.group(1).strip(' -')
    # If text starts with generic facility words, take last 1-4 words; preserve LOC/FRAZ/RIONE contexts
    words=x.split()
    generic={'C.S.','CS','COMUNALE','CENTRO','SPORTIVO','SPORT.','CAMPO','STADIO','ORATORIO','PARROCCHIALE','DI','N.1','N.2','N.3'}
    while words and words[0] in generic: words.pop(0)
    if len(words)<=4:return ' '.join(words)
    # prefer suffix after LOC./FRAZ. if present, otherwise last 2 words
    for marker in ['LOC.','FRAZ.','RIONE']:
        if marker in words:
            i=words.index(marker)
            return ' '.join(words[max(0,i-2):])
    return ' '.join(words[-2:])

# ---------- Simple graphical text calendar ----------
def match_team_line(line, team_names, threshold=.80):
    c=clean(line)
    best=None;score=0
    for t in team_names:
        r=ratio(c,t)
        if r>score:best,score=t,r
    return best if score>=threshold else None

def parse_simple_calendar(text,teams,competition,group):
    y1,y2=season_years(text); phase=text
    team_names=list(teams)
    lines=[x.strip() for x in text.splitlines()]
    matches=[]; i=0
    while i<len(lines):
        ln=clean(lines[i])
        m=re.search(r'\b(\d+)\s*A?\s*GIORNATA\b',ln)
        if not m:
            i+=1;continue
        round_no=m.group(1); i+=1
        # date is next useful line with day/month
        dates=[]
        while i<len(lines) and not dates:
            if re.search(r'\b\d+\s*A?\s*GIORNATA\b',clean(lines[i])):break
            dates=parse_textual_dates(lines[i],y1,y2,phase)
            i+=1
        recog=[]; skip_rest=False
        while i<len(lines):
            cur=clean(lines[i])
            if re.search(r'\b\d+\s*A?\s*GIORNATA\b',cur):break
            if cur.startswith('RIPOSA'):
                skip_rest=True; i+=1; continue
            t=match_team_line(cur,team_names,.82)
            if t:
                if skip_rest:
                    skip_rest=False
                else:
                    recog.append(t)
            i+=1
        for j in range(0,len(recog)-1,2):
            home,away=recog[j],recog[j+1]
            if dates:
                ti=teams.get(home,TeamInfo(home)); matches.append(Match(dates[0],home,away,ti.time,ti.locality,ti.address,round_no))
            if len(dates)>1:
                ti=teams.get(away,TeamInfo(away)); matches.append(Match(dates[1],away,home,ti.time,ti.locality,ti.address,round_no))
    return matches

# ---------- classic ascii calendar ----------
def norm_box_delims(line):
    line=line.replace('!', '|')
    line=re.sub(r'(?<!\S)I(?!\S)', '|', line)
    return line

def extract_box_matches(line):
    # cell contents between visual box delimiters containing a match separator
    line=norm_box_delims(line)
    cells=re.findall(r'\|([^|]+)\|',line)
    out=[]
    for c in cells:
        c=clean(c)
        if 'RIPOSA' in c: continue
        # separator requires whitespace around hyphen to avoid team hyphens
        m=re.match(r'(.+?)\s+-\s+(.+)$',c)
        if m:
            out.append((clean(m.group(1)),clean(m.group(2))))
    return out

def parse_classic_segment(text,teams):
    # Do not parse the field-table portion as calendar rows.
    cut=re.search(r'E\s*L\s*E\s*N\s*C\s*O\s+C\s*A\s*M\s*P\s*I', text, re.I)
    if cut: text=text[:cut.start()]
    lines=text.splitlines(); matches=[]; i=0
    while i<len(lines):
        line=norm_box_delims(lines[i])
        if 'ANDATA:' not in line.upper(): i+=1;continue
        # one or more A/R pairs in the line
        pairs=[]
        pat=re.compile(r'ANDATA:\s*(\d{1,2}/\d{1,2}/\d{2,4})\s*\|.*?RITORNO:\s*(\d{0,2}/?\d{0,2}/?\d{0,4})',re.I)
        for m in pat.finditer(line):
            a=normalize_numeric_date(m.group(1)); r=normalize_numeric_date(m.group(2)) if m.group(2).strip('/') else ''
            pairs.append([a,r,'','',''])
        i+=1
        if i>=len(lines):break
        oreline=norm_box_delims(lines[i])
        # parse each round number and times by cell slices
        # use regex on each GIORNATA occurrence, taking nearest time before and after
        gior=list(re.finditer(r'(\d+)\s*G\s*I\s*O\s*R\s*N\s*A\s*T\s*A',oreline,re.I))
        times_all=[normalize_time(x) for x in re.findall(r'ORE\.*:\s*([0-9:.]*)',oreline,re.I)]
        # Usually 2 times per column
        for idx,p in enumerate(pairs):
            if idx<len(gior): p[2]=gior[idx].group(1)
            if len(times_all)>=2*(idx+1):
                p[3],p[4]=times_all[2*idx],times_all[2*idx+1]
        i+=1
        # skip divider and collect rows until next data line
        while i<len(lines) and 'ANDATA:' not in lines[i].upper():
            if re.match(r'^\s*[.*-]{5,}',lines[i]):
                # may be end block but keep scanning until next date
                i+=1;continue
            rowmatches=extract_box_matches(lines[i])
            for idx,(home,away) in enumerate(rowmatches):
                if idx>=len(pairs):continue
                a,r,rn,ta,tr=pairs[idx]
                # canonicalize names against fields if possible
                home=canonical_team(home,teams) or home; away=canonical_team(away,teams) or away
                ti=teams.get(home,TeamInfo(home))
                matches.append(Match(a,home,away,ti.time or ta,ti.locality,ti.address,rn))
                if r:
                    ti2=teams.get(away,TeamInfo(away))
                    matches.append(Match(r,away,home,ti2.time or tr,ti2.locality,ti2.address,rn))
            i+=1
    return matches

def canonical_team(name,teams,threshold=.78):
    if name in teams:return name
    best=None;sc=0
    for t in teams:
        r=ratio(name,t)
        if r>sc:best,sc=t,r
    return best if sc>=threshold else None

# ---------- graphical OCR helpers ----------
def render_page(page,zoom=2):
    pix=page.get_pixmap(matrix=fitz.Matrix(zoom,zoom),alpha=False)
    img=np.frombuffer(pix.samples,dtype=np.uint8).reshape(pix.height,pix.width,pix.n)
    if pix.n==4:return cv2.cvtColor(img,cv2.COLOR_RGBA2BGR)
    return cv2.cvtColor(img,cv2.COLOR_RGB2BGR)

def ocr_img(img,psm=6,whitelist=None):
    if img.size==0:return ''
    g=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    g=cv2.resize(g,None,fx=2.5,fy=2.5,interpolation=cv2.INTER_CUBIC)
    _,th=cv2.threshold(g,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    cfg=f'--oem 3 --psm {psm}'
    if whitelist:cfg+=f' -c tessedit_char_whitelist={whitelist}'
    try: txt=pytesseract.image_to_string(th,lang='ita',config=cfg)
    except: txt=pytesseract.image_to_string(th,config=cfg)
    return clean(txt)

def direct_text_quality(text):
    if not text:return 0
    good=sum(ch.isascii() and (ch.isalnum() or ch.isspace() or ch in "-:./'()") for ch in text)
    return good/max(1,len(text))

def find_white_calendar_boxes(img):
    H,W=img.shape[:2]; gray=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    mask=cv2.inRange(gray,218,255)
    mask=cv2.morphologyEx(mask,cv2.MORPH_CLOSE,cv2.getStructuringElement(cv2.MORPH_RECT,(5,5)))
    cnts,_=cv2.findContours(mask,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
    rects=[]
    for c in cnts:
        x,y,w,h=cv2.boundingRect(c)
        if .14*W<w<.27*W and .07*H<h<.22*H and .15*H<y<.92*H:
            rects.append((x,y,w,h))
    # cluster/sort row then x
    rects=sorted(rects,key=lambda r:(round(r[1]/max(1,H*.04)),r[0]))
    return rects

def longest_regular_run(vals):
    vals=sorted(set(vals))
    if len(vals)<3:return vals
    diffs=[b-a for a,b in zip(vals,vals[1:]) if 10<=b-a<=80]
    if not diffs:return vals
    med=statistics.median(diffs)
    best=[];cur=[vals[0]]
    for a,b in zip(vals,vals[1:]):
        if abs((b-a)-med)<=4:
            cur.append(b)
        else:
            if len(cur)>len(best):best=cur
            cur=[b]
    if len(cur)>len(best):best=cur
    return best

def parse_graphic_field_table(page):
    """Parse modern graphical field table using dynamically detected row/column lines."""
    img=render_page(page,2);H,W=img.shape[:2];gray=cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
    bw=cv2.threshold(gray,205,255,cv2.THRESH_BINARY_INV)[1]
    kh=cv2.getStructuringElement(cv2.MORPH_RECT,(max(40,W//20),1)); hl=cv2.morphologyEx(bw,cv2.MORPH_OPEN,kh)
    cnts,_=cv2.findContours(hl,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE); ys=[]
    for c in cnts:
        x,y,w,h=cv2.boundingRect(c)
        if w>.65*W and .12*H<y<.9*H:ys.append(y)
    ys=longest_regular_run(ys)
    if len(ys)<8:return {}
    gap=int(round(statistics.median([b-a for a,b in zip(ys,ys[1:])])))
    rowbounds=[ys[0]-gap]+ys
    kv=cv2.getStructuringElement(cv2.MORPH_RECT,(1,max(40,H//15)));vl=cv2.morphologyEx(bw,cv2.MORPH_OPEN,kv)
    cnts,_=cv2.findContours(vl,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE);xs=[]
    for c in cnts:
        x,y,w,h=cv2.boundingRect(c)
        if h>.3*H and .1*H<y<.9*H:xs.append(x)
    xs=sorted(set(xs))
    if len(xs)<6:return {}
    xs=xs[:6] if len(xs)==6 else xs[-6:]
    teams={}
    for ya,yb in zip(rowbounds,rowbounds[1:]):
        if ya<0:continue
        # One OCR call per complete row: much faster on Streamlit Cloud.
        row=img[ya+1:yb-1,xs[0]:xs[-1]]
        g=cv2.cvtColor(row,cv2.COLOR_BGR2GRAY)
        g=cv2.resize(g,None,fx=3,fy=3,interpolation=cv2.INTER_CUBIC)
        _,th=cv2.threshold(g,180,255,cv2.THRESH_BINARY)
        try: raw=pytesseract.image_to_string(th,lang='ita',config='--psm 7')
        except: raw=pytesseract.image_to_string(th,config='--psm 7')
        parts=[clean(x) for x in raw.split('|') if clean(x)]
        if len(parts)<3:continue
        # Usually: name | code | field/locality | address | time
        code_idx=None
        for j,x in enumerate(parts):
            if re.fullmatch(r'\d{1,5}',x):code_idx=j;break
        if code_idx is None or code_idx==0:continue
        name=parts[code_idx-1]
        fieldloc=parts[code_idx+1] if code_idx+1<len(parts) else ''
        rest=parts[code_idx+2:]
        tm='';addr=''
        for x in rest:
            nt=normalize_time(x)
            if nt:tm=nt
            elif not addr:addr=x
            else:addr+=' '+x
        name=re.sub(r'^[^A-Z0-9]+','',name).strip()
        if len(name)<3 or name.startswith('SOCIETA'):continue
        locality=extract_locality_from_field(fieldloc)
        teams[name]=TeamInfo(name,locality,addr,tm,fieldloc)
    return teams




def _group_word_lines(words, tol=2.2):
    """Raggruppa le parole PDF che appartengono alla stessa riga visiva."""
    rows=[]
    for w in sorted(words,key=lambda x:((x[1]+x[3])/2,x[0])):
        yc=(w[1]+w[3])/2
        if not rows or abs(yc-rows[-1][0])>tol:
            rows.append([yc,[w]])
        else:
            rows[-1][1].append(w)
    return rows


def _team_base_key(name):
    """Chiave squadra senza le sole forme societarie finali (ASD, SSD, SRL...)."""
    s=clean(name)
    patterns=[
        r'S\.?S\.?D\.?\s*A\.?\s*R\.?\s*L\.?$', r'SSD\s*A\s*R\s*L$',
        r'S\.?S\.?D\.?$', r'SSD(?:ARL|RL)?$', r'A\.?S\.?D\.?$', r'ASD$',
        r'S\.?R\.?L\.?$', r'SRL$', r'A\.?\s*R\.?\s*L\.?$', r'ARL$',
        r'C\.?V\.?$', r'F\.?B\.?C\.?$', r'POL\.?D\.?$', r'POL\.?$'
    ]
    changed=True
    while changed:
        changed=False
        for pat in patterns:
            ns=re.sub(r'\s*\b'+pat,'',s).strip(' .-')
            if ns!=s:
                s=ns; changed=True
    return re.sub(r'[^A-Z0-9]+','',s)


def smart_canonical_team(name, teams):
    """Associa il nome abbreviato del calendario alla riga corretta della tabella campi."""
    c=clean(name)
    # Nei calendari CRL "AC." / "ACC." è spesso abbreviazione di ACADEMY/ACCADEMIA.
    ab=re.match(r'^(?:ACC?\.)\s*(.+)$',c)
    if ab:
        rem=_team_base_key(ab.group(1)); candidates=[]
        for t in teams:
            mt=re.match(r'^(?:ACADEMY|ACCADEMIA)\s+(.+)$',clean(t))
            if mt:
                sc=SequenceMatcher(None,rem,_team_base_key(mt.group(1))).ratio()
                candidates.append((sc,t))
        if candidates:
            sc,t=max(candidates)
            if sc>=.72:
                return t

    q=_team_base_key(name)
    exact=[t for t in teams if _team_base_key(t)==q]
    if len(exact)==1:
        return exact[0]

    best=None; best_score=0
    for t in teams:
        k=_team_base_key(t)
        sc=SequenceMatcher(None,q,k).ratio()
        if len(q)>=5 and (q in k or k in q):
            extra=abs(len(q)-len(k))
            sc=max(sc,.93-min(.20,extra*.01))
        if sc>best_score:
            best,best_score=t,sc
    return best if best_score>=.66 else None


def parse_modern_field_table_words_v2(page):
    """
    Legge le nuove tabelle CRL/LND con intestazioni:
    Società | N. | Campo / Località | Indirizzo | Orario | Giorno.
    Usa le coordinate delle parole del PDF, non OCR e non coordinate verticali fisse.
    """
    words=page.get_text('words') or []
    def arr(pred): return [w for w in words if pred(clean(w[4]))]
    soc=arr(lambda t:t.startswith('SOCIETA'))
    code=arr(lambda t:t in {'N.','N'})
    campo=arr(lambda t:t=='CAMPO')
    addr=arr(lambda t:t.startswith('INDIRIZZO'))
    tm=arr(lambda t:t.startswith('ORARIO'))
    day=arr(lambda t:t.startswith('GIORNO'))
    if not (soc and code and campo and addr):
        return {}

    hdr_y=min(w[1] for w in soc+code+campo+addr)
    pick=lambda a:min(a,key=lambda w:abs(w[1]-hdr_y))
    sx,cx,fx,ax=pick(soc)[0],pick(code)[0],pick(campo)[0],pick(addr)[0]
    tx=pick(tm)[0] if tm else page.rect.width
    dx=pick(day)[0] if day else page.rect.width

    data=[w for w in words if w[1]>hdr_y+7 and w[1]<page.rect.height*.90]
    teams={}
    for _,ws in _group_word_lines(data,2.2):
        ordered=sorted(ws,key=lambda q:q[0])
        code_txt=clean(' '.join(w[4] for w in ordered if cx-4<=w[0]<fx-4))
        if not re.fullmatch(r'\d{1,5}',code_txt):
            continue
        name=clean(' '.join(w[4] for w in ordered if sx-4<=w[0]<cx-4))
        fieldloc=clean(' '.join(w[4] for w in ordered if fx-4<=w[0]<ax-4))
        if tm:
            address=clean(' '.join(w[4] for w in ordered if ax-4<=w[0]<tx-4))
            time_txt=clean(' '.join(w[4] for w in ordered if tx-4<=w[0]<dx-4))
        else:
            address=clean(' '.join(w[4] for w in ordered if ax-4<=w[0]<page.rect.width))
            time_txt=''
        if not name or name.startswith('SOCIETA'):
            continue
        teams[name]=TeamInfo(
            name=name,
            locality=extract_locality_from_field(fieldloc),
            address=address,
            time=normalize_time(time_txt),
            field_name=fieldloc,
        )
    return teams


def detect_modern_round_headers(page):
    """Trova GIORNATA 1..N tramite posizione delle parole nel PDF."""
    words=page.get_text('words') or []
    headers=[]
    for w in words:
        if clean(w[4])!='GIORNATA':
            continue
        yc=(w[1]+w[3])/2
        candidates=[]
        for q in words:
            if not re.fullmatch(r'\d{1,2}',clean(q[4])):
                continue
            qy=(q[1]+q[3])/2
            if abs(qy-yc)<3 and q[0]>=w[2]-3 and q[0]-w[2]<30:
                candidates.append(q)
        if candidates:
            q=min(candidates,key=lambda z:z[0])
            headers.append({'rn':int(q[4]),'x':w[0],'y':w[1]})
    # elimina eventuali duplicati mantenendo una sola intestazione per numero
    return list({h['rn']:h for h in headers}.values())


def parse_modern_layout_calendar(page, teams):
    """
    Parser dei calendari grafici 2026/27 (Eccellenza, U14/U15/U16/U17 ecc.).
    La pagina contiene più riquadri affiancati: il normale get_text() mescola le
    colonne. Qui ricostruiamo ogni riquadro dalle coordinate X/Y delle parole.
    """
    words=page.get_text('words') or []
    headers=detect_modern_round_headers(page)
    if len(headers)<3:
        return []

    header_rows=[]
    for h in sorted(headers,key=lambda z:(z['y'],z['x'])):
        if not header_rows or abs(h['y']-header_rows[-1][0])>8:
            header_rows.append([h['y'],[h]])
        else:
            header_rows[-1][1].append(h)

    matches=[]
    for ri,(hy,group) in enumerate(header_rows):
        group=sorted(group,key=lambda z:z['x'])
        slots=len(group)
        next_hy=header_rows[ri+1][0] if ri+1<len(header_rows) else page.rect.height*.86
        band=[w for w in words if w[1]>=hy-2 and w[1]<next_hy-4]

        # Le righe gara hanno un trattino centrale per ciascun riquadro.
        match_lines=[]
        for yc,lw in _group_word_lines(band,2.0):
            seps=sorted([w for w in lw if clean(w[4]) in {'-','–','—'}],key=lambda w:w[0])
            if len(seps)>=slots and yc>hy+20:
                match_lines.append((yc,sorted(lw,key=lambda w:w[0]),seps))
        if not match_lines:
            continue

        # Centro di ciascun riquadro = posizione mediana del trattino CASA-OSPITE.
        centers=[]
        for j in range(slots):
            vals=[(seps[j][0]+seps[j][2])/2 for _,_,seps in match_lines if len(seps)>=slots]
            centers.append(statistics.median(vals))
        bounds=[0]+[(centers[j]+centers[j+1])/2 for j in range(slots-1)]+[page.rect.width]

        # Date A/R: assegna ciascuna data al riquadro più vicino.
        first_match_y=min(y for y,_,_ in match_lines)
        date_map={j:[] for j in range(slots)}
        for w in band:
            if w[1]>=first_match_y:
                continue
            dt=normalize_numeric_date(w[4])
            if not dt:
                continue
            xc=(w[0]+w[2])/2
            j=min(range(slots),key=lambda k:abs(xc-centers[k]))
            if dt not in date_map[j]:
                date_map[j].append(dt)

        pair_map={j:[] for j in range(slots)}
        for _,lw,seps in match_lines:
            for j,c in enumerate(centers):
                sep=min(seps,key=lambda z:abs(((z[0]+z[2])/2)-c))
                sc=(sep[0]+sep[2])/2
                x0,x1=bounds[j],bounds[j+1]
                left=[w for w in lw if x0<=((w[0]+w[2])/2)<sc-1 and clean(w[4]) not in {'-','–','—'}]
                right=[w for w in lw if sc+1<((w[0]+w[2])/2)<x1 and clean(w[4]) not in {'-','–','—'}]
                home_raw=clean(' '.join(w[4] for w in sorted(left,key=lambda w:w[0])))
                away_raw=clean(' '.join(w[4] for w in sorted(right,key=lambda w:w[0])))
                if home_raw and away_raw:
                    pair_map[j].append((home_raw,away_raw))

        for j,h in enumerate(group):
            dates=date_map[j]
            if not dates:
                continue
            andata=dates[0]
            ritorno=dates[1] if len(dates)>1 else ''
            for home_raw,away_raw in pair_map[j]:
                home=smart_canonical_team(home_raw,teams) or home_raw
                away=smart_canonical_team(away_raw,teams) or away_raw
                ti=teams.get(home,TeamInfo(home))
                matches.append(Match(andata,home,away,ti.time,ti.locality,ti.address,str(h['rn'])))
                if ritorno:
                    ti2=teams.get(away,TeamInfo(away))
                    matches.append(Match(ritorno,away,home,ti2.time,ti2.locality,ti2.address,str(h['rn'])))
    return matches

def parse_modern_native_calendar(page, teams):
    """
    Legge i calendari grafici moderni direttamente dal testo nativo del PDF.

    In molti PDF CRL/LND l'ordine interno è:
        CASA
        -
        OSPITE
        ... (tutte le gare della giornata)
        GIORNATA N
        A. gg/mm/aaaa
        R. gg/mm/aaaa

    È molto più stabile dell'OCR su Streamlit Cloud e impedisce che una
    singola gara venga persa per differenze di versione di Tesseract.
    """
    raw = page.get_text('text') or ''
    lines = [clean(x) for x in raw.splitlines() if clean(x)]
    if not lines:
        return []

    # Occorrenze delle intestazioni GIORNATA nel testo nativo.
    headers = []
    for i, ln in enumerate(lines):
        m = re.search(r'GIORNATA\s*(\d+)|(?:^|\s)(\d+)\s*GIORNATA', ln, re.I)
        if m:
            rn = m.group(1) or m.group(2)
            headers.append((i, rn))
    if not headers:
        return []

    out = []
    segment_start = 0
    parsed_rounds = 0

    for hidx, (i, round_no) in enumerate(headers):
        # Le gare della giornata precedono l'intestazione GIORNATA.
        segment = lines[segment_start:i]
        pairs = []

        # Formato più comune: CASA / '-' / OSPITE.
        for j in range(1, len(segment)-1):
            sep = segment[j].strip()
            if sep not in {'-', '–', '—', '='}:
                continue
            home_raw = segment[j-1]
            away_raw = segment[j+1]
            if not home_raw or not away_raw:
                continue
            home = canonical_team(home_raw, teams, .62)
            away = canonical_team(away_raw, teams, .62)
            if home and away and home != away:
                pair = (home, away)
                if pair not in pairs:
                    pairs.append(pair)

        # Fallback: alcune estrazioni native tengono "CASA - OSPITE" su una riga.
        if not pairs:
            for ln in segment:
                m = re.match(r'(.+?)\s+[-–—]\s+(.+)$', ln)
                if not m:
                    continue
                home = canonical_team(m.group(1), teams, .62)
                away = canonical_team(m.group(2), teams, .62)
                if home and away and home != away:
                    pair = (home, away)
                    if pair not in pairs:
                        pairs.append(pair)

        # Date A./R. sono immediatamente dopo l'intestazione.
        da = ''
        dr = ''
        scan_end = headers[hidx+1][0] if hidx+1 < len(headers) else min(len(lines), i+8)
        # Non serve arrivare alla giornata successiva: bastano poche righe.
        scan_end = min(scan_end, i+8)
        for ln in lines[i+1:scan_end]:
            ma = re.search(r'\bA[.,]?\s*(\d{1,2}[./-]\d{1,2}[./-]20\d{2})', ln, re.I)
            mr = re.search(r'\bR[.,]?\s*(\d{1,2}[./-]\d{1,2}[./-]20\d{2})', ln, re.I)
            if ma and not da:
                da = normalize_numeric_date(ma.group(1))
            if mr and not dr:
                dr = normalize_numeric_date(mr.group(1))

        if pairs and da:
            parsed_rounds += 1
            for home, away in pairs:
                ti = teams.get(home, TeamInfo(home))
                out.append(Match(da, home, away, ti.time, ti.locality, ti.address, str(round_no)))
                if dr:
                    ti2 = teams.get(away, TeamInfo(away))
                    out.append(Match(dr, away, home, ti2.time, ti2.locality, ti2.address, str(round_no)))

        # Il segmento successivo parte dopo le righe A./R.
        next_start = i + 1
        while next_start < len(lines) and next_start < i + 8:
            ln = lines[next_start]
            if re.search(r'\b[AR][.,]?\s*\d{1,2}[./-]\d{1,2}[./-]20\d{2}', ln, re.I):
                next_start += 1
                continue
            break
        segment_start = next_start

    # Usa il parser nativo solo se ha letto praticamente tutte le giornate.
    # Altrimenti si lascia lavorare il fallback OCR già esistente.
    if parsed_rounds >= max(1, int(len(headers) * .85)):
        return out
    return []

def parse_graphic_calendar(page,teams):
    # Prima prova il testo nativo: è deterministico e non dipende dalla
    # versione di Tesseract installata su Streamlit Cloud.
    native = parse_modern_native_calendar(page, teams)
    if native:
        return native

    img=render_page(page,2)
    boxes=find_white_calendar_boxes(img)

    # --------------------------------------------------------
    # DATE DELLE GIORNATE
    # --------------------------------------------------------
    # Nei PDF grafici moderni il testo delle squadre può avere una
    # mappa-font problematica, mentre le date A./R. sono spesso
    # perfettamente estraibili dal testo nativo del PDF.
    # Usiamo quindi PRIMA il testo nativo e OCR solo come fallback.
    # Questo evita, per esempio, di perdere il ritorno dell'ultima
    # giornata quando Tesseract non riconosce una singola data.
    direct_text=page.get_text('text') or ''

    a_direct=re.findall(
        r'\bA[.,]?\s*(\d{1,2}/\d{1,2}/20\d{2})',
        direct_text,
        re.I
    )
    r_direct=re.findall(
        r'\bR[.,]?\s*(\d{1,2}/\d{1,2}/20\d{2})',
        direct_text,
        re.I
    )

    a_dates=[normalize_numeric_date(x) for x in a_direct]
    r_dates=[normalize_numeric_date(x) for x in r_direct]
    a_dates=[x for x in a_dates if x]
    r_dates=[x for x in r_dates if x]

    # Se il testo nativo non restituisce abbastanza date, completa
    # la lettura tramite OCR dell'intera pagina.
    if len(a_dates)<len(boxes) or len(r_dates)<len(boxes):
        whole=ocr_img(img,6)
        a_ocr=[normalize_numeric_date(x) for x in re.findall(
            r'\bA[.,]?\s*(\d{1,2}/\d{1,2}/20\d{2})',
            whole,
            re.I
        )]
        r_ocr=[normalize_numeric_date(x) for x in re.findall(
            r'\bR[.,]?\s*(\d{1,2}/\d{1,2}/20\d{2})',
            whole,
            re.I
        )]
        a_ocr=[x for x in a_ocr if x]
        r_ocr=[x for x in r_ocr if x]

        # Preferisce la fonte che ha riconosciuto più giornate.
        if len(a_ocr)>len(a_dates):
            a_dates=a_ocr
        if len(r_ocr)>len(r_dates):
            r_dates=r_ocr

    names=list(teams)
    out=[]
    for idx,box in enumerate(boxes):
        x,y,w,h=box
        txt=ocr_img(img[y:y+h,x:x+w],6)
        lines=[clean(z) for z in txt.split(' | ') if clean(z)] if ' | ' in txt else [clean(z) for z in re.split(r'[\n\r]+',pytesseract.image_to_string(cv2.resize(cv2.cvtColor(img[y:y+h,x:x+w],cv2.COLOR_BGR2GRAY),None,fx=3,fy=3),config='--psm 6')) if clean(z)]
        pairs=[]
        for ln in lines:
            # identify two best teams in line by fuzzy substring/ratio on left-right separator candidates
            # separators OCR may be -, +, =, —
            cand_parts=re.split(r'\s+[-+—=]+\s+',ln,maxsplit=1)
            if len(cand_parts)==2:
                hname=canonical_team(cand_parts[0],teams,.62); aname=canonical_team(cand_parts[1],teams,.62)
                if hname and aname and hname!=aname:pairs.append((hname,aname));continue
            # fallback: try every pair and compare concatenation
            best=None;sc=0
            for hname in names:
                for aname in names:
                    if hname==aname:continue
                    rr=ratio(ln,hname+' '+aname)
                    if rr>sc:best,sc=(hname,aname),rr
            if best and sc>.72:pairs.append(best)
        # dedupe preserve order
        uniq=[]
        for p in pairs:
            if p not in uniq:uniq.append(p)
        da=a_dates[idx] if idx<len(a_dates) else ''
        dr=r_dates[idx] if idx<len(r_dates) else ''
        for home,away in uniq:
            ti=teams.get(home,TeamInfo(home));out.append(Match(da,home,away,ti.time,ti.locality,ti.address,str(idx+1)))
            if dr:
                ti2=teams.get(away,TeamInfo(away));out.append(Match(dr,away,home,ti2.time,ti2.locality,ti2.address,str(idx+1)))
    return out

# ---------- Programma gare ----------
def parse_programma_gare_page(text):
    lines=[clean(x) for x in text.splitlines() if clean(x)]
    comp='COPPA ITALIA ECCELLENZA'
    for ln in lines[:8]:
        if 'CAMPIONATO' in ln:
            comp=re.sub(r'^CAMPIONATO\s+[A-Z]{1,3}\s*','',ln).strip() or comp
    sections=[]; i=0
    while i<len(lines):
        gm=re.match(r'GIRONE\s+(\d+|[A-Z])$',lines[i])
        if not gm:i+=1;continue
        group=gm.group(1); start=i+1; i+=1
        while i<len(lines) and not re.match(r'GIRONE\s+(\d+|[A-Z])$',lines[i]):i+=1
        seg=lines[start:i]
        # remove headers
        seg=[x for x in seg if x not in ['DATA','ORA','PROGRAMMA GARE']]
        matches=[]; last=0
        for di,s in enumerate(seg):
            if re.fullmatch(r'\d{1,2}/\d{1,2}/20\d{2}',s):
                if di>=3 and di+2<len(seg):
                    home,away,venue=seg[di-3],seg[di-2],seg[di-1]
                    mt=re.match(r'([0-2]?\d[:.]\d{2})\s+\w+\s+(.+)$',seg[di+1])
                    if mt:
                        tm=normalize_time(mt.group(1)); loc=clean(mt.group(2)); addr=seg[di+2]
                        matches.append(Match(normalize_numeric_date(s),home,away,tm,loc,addr,''))
        if matches:
            teams={}
            for m in matches:
                teams.setdefault(m.home,TeamInfo(m.home,m.locality,m.address,m.time, ''))
                teams.setdefault(m.away,TeamInfo(m.away))
            sections.append(Section(comp,group,teams,matches,'programma_gare'))
    return sections

# ---------- document-level ----------
def parse_pdf(path):
    doc=fitz.open(path)
    texts=[p.get_text('text') for p in doc]
    alltext='\n'.join(texts)

    # Programma Gare / Coppe: struttura tabellare diversa dal calendario A/R.
    if 'PROGRAMMA GARE' in alltext.upper() and re.search(r'GIRONE\s+\d+',alltext,re.I):
        secs=[]
        for t in texts:
            secs.extend(parse_programma_gare_page(t))
        return secs

    sections=[]
    used=set()

    # 1) Nuovi calendari grafici CRL/LND: pagina calendario + pagina tabella campi.
    #    Non cerchiamo la frase "ELENCO CAMPI": nei file U14/U15/U16/U17 non c'è.
    for pi in range(len(doc)-1):
        if pi in used:
            continue
        headers=detect_modern_round_headers(doc[pi])
        teams_modern=parse_modern_field_table_words_v2(doc[pi+1])
        if len(headers)>=3 and len(teams_modern)>=4:
            comp,group=parse_header_comp_group(texts[pi+1])
            matches=parse_modern_layout_calendar(doc[pi],teams_modern)
            if matches:
                sections.append(Section(comp or 'CALENDARIO',group,teams_modern,matches,'modern_native_layout'))
                used.update([pi,pi+1])

    # 2) Formati classici / provinciali già supportati.
    for pi,t in enumerate(texts):
        if pi in used:
            continue
        nxt=texts[pi+1] if pi+1<len(texts) else ''
        if ('GIORNATA' in t.upper() or 'G I O R N A T A' in t.upper()) and ('E L E N C O' in nxt.upper() and ('CAMPI' in nxt.upper() or 'C A M P I' in nxt.upper())):
            comp,group=parse_header_comp_group(t+'\n'+nxt)
            teams=parse_field_table_words(doc[pi+1])
            if len(teams)<5:
                teams=parse_field_table_text(nxt)
            if len(teams)<5:
                teams=parse_graphic_field_table(doc[pi+1])
            if 'ANDATA:' in t.upper():
                matches=parse_classic_segment(t,teams); fmt='classic'
            elif direct_text_quality(t)>.80:
                matches=parse_simple_calendar(t,teams,comp,group); fmt='simple_graphic_text'
            else:
                matches=parse_graphic_calendar(doc[pi],teams); fmt='graphic_ocr'
            if matches:
                sections.append(Section(comp or 'CALENDARIO',group,teams,matches,fmt))
                used.update([pi,pi+1])

    # 3) Fallback OCR per vecchi PDF grafici con mappa-font corrotta.
    if not sections and len(doc)>=2:
        teams=parse_graphic_field_table(doc[1])
        if teams:
            ocrhead=ocr_img(render_page(doc[1],2)[:380,:,:],6)
            comp,group=parse_header_comp_group(ocrhead)
            first_clean=''
            for ln in texts[1].splitlines():
                c=clean(ln)
                if c and sum(ch.isascii() for ch in c)/max(1,len(c))>.9 and len(c)>=5:
                    first_clean=c
                    break
            if first_clean and 'GIRONE' not in first_clean and len(first_clean)<80:
                comp=first_clean
            matches=parse_graphic_calendar(doc[0],teams)
            if matches:
                sections.append(Section(comp or 'CALENDARIO',group,teams,matches,'graphic_ocr'))

    return [s for s in sections if s.matches]

def parse_docx(path):
    from docx import Document
    d=Document(path)
    text='\n'.join(p.text for p in d.paragraphs)
    # split by committee section header, retaining each segment
    starts=[m.start() for m in re.finditer(r'\*\s*[A-Z0-9 .\-]+\s+GIRONE:\s*[A-Z0-9]+\s*\*',text,re.I)]
    if not starts:starts=[0]
    starts.append(len(text));secs=[]
    for a,b in zip(starts,starts[1:]):
        seg=text[a:b]
        comp,group=parse_header_comp_group(seg)
        teams=parse_field_table_text(seg)
        matches=parse_classic_segment(seg,teams)
        if matches:secs.append(Section(comp or 'CALENDARIO',group,teams,matches,'classic_docx'))
    return secs


# ============================================================
# STREAMLIT APP
# ============================================================

def indirizzo_excel(match):
    addr=clean(match.address)
    loc=clean(match.locality)
    if addr and loc:
        return f"{addr} - {loc}"
    return addr or loc


def sort_date_value(s):
    try:
        return datetime.strptime(s, '%d/%m/%Y')
    except:
        return datetime.max


def create_excel_for_team(section, selected_team):
    selected=[m for m in section.matches if m.home==selected_team or m.away==selected_team]
    selected=sorted(selected,key=lambda m:(sort_date_value(m.date),m.time,m.home,m.away))

    wb=Workbook()
    ws=wb.active
    ws.title='Calendario'
    ws.append(['Data','Ora','Tipo','Squadra casa','Squadra ospite','Indirizzo'])
    for m in selected:
        ws.append([
            m.date,
            m.time,
            'CAMPIONATO',
            m.home,
            m.away,
            indirizzo_excel(m),
        ])

    for c in ws[1]:
        c.font=Font(bold=True)
        c.alignment=Alignment(horizontal='center')
    widths=[14,10,16,32,32,52]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes='A2'
    ws.auto_filter.ref=ws.dimensions

    out=io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue(), selected


def safe_filename(s):
    s=clean(s)
    s=re.sub(r'[^A-Z0-9]+','_',s).strip('_')
    return s or 'SQUADRA'


@st.cache_data(show_spinner=False)
def analyze_upload_v2(file_bytes, filename):
    suffix=Path(filename).suffix.lower()
    if suffix not in ['.pdf','.docx']:
        raise ValueError('Sono supportati file PDF e DOCX.')
    tmp=None
    try:
        with tempfile.NamedTemporaryFile(delete=False,suffix=suffix) as f:
            f.write(file_bytes)
            tmp=f.name
        if suffix=='.docx':
            return parse_docx(tmp)
        return parse_pdf(tmp)
    finally:
        if tmp and os.path.exists(tmp):
            try: os.remove(tmp)
            except: pass


st.set_page_config(page_title='Calendario → Excel', page_icon='⚽', layout='centered')
st.title('⚽ Calendario → Excel')
st.caption(f'Versione app: {PARSER_VERSION}')
st.write('Carica un calendario LND/FIGC, scegli il girone e la squadra, quindi scarica l’Excel.')

uploaded=st.file_uploader('1. Carica il calendario', type=['pdf','docx'])

if uploaded is not None:
    with st.spinner('Analisi del calendario in corso…'):
        try:
            sections=analyze_upload_v2(uploaded.getvalue(), uploaded.name)
        except Exception as e:
            st.error(f'Errore durante la lettura del file: {e}')
            st.stop()

    if not sections:
        st.error('Non sono riuscito a riconoscere partite nel file. Il formato potrebbe essere diverso da quelli attualmente supportati.')
        st.stop()

    st.success(f'Analisi completata: {len(sections)} sezione/i di calendario riconosciuta/e.')

    labels=[]
    for idx,s in enumerate(sections,1):
        label=s.label or f'Sezione {idx}'
        # Make duplicated labels distinguishable.
        if label in labels:
            label=f'{label} ({idx})'
        labels.append(label)

    if len(sections)>1:
        chosen_label=st.selectbox('2. Seleziona categoria / girone', labels)
        section=sections[labels.index(chosen_label)]
    else:
        section=sections[0]
        st.info(f'Categoria/Girone: {section.label}')

    teams=sorted(set([m.home for m in section.matches]+[m.away for m in section.matches]))
    if not teams:
        st.error('Nessuna squadra riconosciuta nella sezione selezionata.')
        st.stop()

    selected_team=st.selectbox('3. Per quale squadra vuoi l\'estrapolazione?', teams)
    team_matches=[m for m in section.matches if m.home==selected_team or m.away==selected_team]
    st.write(f'Partite trovate per **{selected_team}**: **{len(team_matches)}**')

    with st.expander('Dettagli analisi'):
        st.write(f'Versione parser: `{PARSER_VERSION}`')
        st.write(f'Formato riconosciuto: `{section.source_format}`')
        st.write(f'Squadre nel girone: {len(teams)}')
        st.write(f'Partite complessive lette: {len(section.matches)}')

    if st.button('4. Genera Excel', type='primary', use_container_width=True):
        excel_bytes, extracted=create_excel_for_team(section,selected_team)
        st.session_state['excel_bytes']=excel_bytes
        st.session_state['excel_name']=f'Calendario_{safe_filename(selected_team)}.xlsx'
        st.session_state['excel_count']=len(extracted)
        st.session_state['excel_key']=(uploaded.name,section.label,selected_team)

    current_key=(uploaded.name,section.label,selected_team)
    if st.session_state.get('excel_key')==current_key and 'excel_bytes' in st.session_state:
        st.success(f"Excel pronto: {st.session_state.get('excel_count',0)} partite estratte.")
        st.download_button(
            '⬇️ Scarica Excel',
            data=st.session_state['excel_bytes'],
            file_name=st.session_state['excel_name'],
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True,
        )

st.divider()
st.caption('Colonne Excel: Data | Ora | CAMPIONATO | Squadra casa | Squadra ospite | Indirizzo - Paese')
